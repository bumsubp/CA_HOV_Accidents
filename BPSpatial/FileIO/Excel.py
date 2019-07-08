import xlrd
import openpyxl
import csv
import xlsxwriter

import BPSpatial.Constants as C

def clear_sheet(filepath, sheetname):
    """
    Method for deleting all the values in a excel sheet
    
    Parameters
    ----------
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
    """
    wb = openpyxl.load_workbook(filepath)
    std = wb[sheetname]
    wb.remove(std)
    wb.create_sheet(sheetname)
    wb.save(filepath)
    print("Successfully cleared!")
    
def read_csv(filepath):
    """
    Method for reading a .csv file which separating values with comma(' ,')
    
    Parameter
    ---------
    filepath:   string
                path of the .csv file to handle
    
    Return
    ------
    resultList: nested list
                list of values of .csv file \n
                [[row1val1, row1val2, ...], \n
                [row2val1, row2val2, ...], ...]
    """
    with open(filepath, newline='') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
        
        outList = []
        for row in spamreader:
            outList.append(row)
        return outList
    
def create_xlsx(filepath, sheetname):
    """
    Method for creating a bland excel file
    
    Parameters
    ----------
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
    """
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook(filepath)
    workbook.add_worksheet(sheetname)
    workbook.close()
    print("Successfully created!")
    
def excel_to_csv(filepath, sheetname, csvfilepath, encoding='utf-8'):
    """
    Method for converting excel file to csv file
    
    Parameters
    ----------
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
    
    csvfilepath:string
                path of csv file to save
                
    encoding:   string
    """
    wb = xlrd.open_workbook(filepath)
    sh = wb.sheet_by_name(sheetname)
    csvfile= open(csvfilepath, 'w', newline='')
    wr = csv.writer(csvfile, quoting=csv.QUOTE_ALL)

    for rownum in range(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    csvfile.close()
    print("Successfully converted!")


def read_xlsx(filepath, sheetname):
    """
    Method for loading an excel file
    
    Parameters
    ----------
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
                
    Return
    ------
    resultXlrd: list of xlrd.sheet.Cell
    
                list of excel values in xlrd.sheet.Cell type
                
                value.ctype \n
                XL_CELL_EMPTY	0	empty string '' \n
                XL_CELL_TEXT	1	a Unicode string \n
                XL_CELL_NUMBER	2	float \n
                XL_CELL_DATE	3	float \n
                XL_CELL_BOOLEAN	4	int; 1 means TRUE, 0 means FALSE \n
                XL_CELL_ERROR	5	int representing internal Excel codes; 
                for a text representation, refer to the supplied dictionary error_text_from_code \n
                XL_CELL_BLANK	6	empty string ''. Note: this type will appear only when open_workbook(..., formatting_info=True) is used. \n
                
                value.value \n
                Or output of this function can be an input of 'excel_to_list' function
                
    """
    workbook = xlrd.open_workbook(filepath)
    worksheet = workbook.sheet_by_name(sheetname)
    
    num_rows = worksheet.nrows -1
    curr_row = -1
    resultXlrd = []
    
    while curr_row < num_rows:
        curr_row += 1
        row = worksheet.row(curr_row)
        resultXlrd.append(row)
    
    return resultXlrd
    

def excel_to_list(excelList, startAt=0):
    """
    Method for coverting list of xlrd.sheet.Cell to a nested list
    
    Parameters
    ----------
    excelList:  list of xlrd.sheet.Cell
                usually output of read_xlsx function
                
    startAt:    int
                start index of rows \n
                if you have a header in the excel file, this would be 1
                
    Return
    ------
    resultList: nested list
                list of values of .csv file \n
                [[row1val1, row1val2, ...], \n
                [row2val1, row2val2, ...], ...]
    """
    resultList = []
    for row in excelList[startAt:]:
        rowList = []
        for record in row:
            if record.ctype == 5:
                rowList.append('N/A')
            else:
                rowList.append(record.value)
        resultList.append(rowList)
    return resultList

    
def write_column(filepath, sheetname, insert, startCol=0, startRow=1):
    """
    Method for adding a set of columns in an existing excel file
    
    Paramters
    ---------
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
        
    insert:     nested list
                [ [col1], [col2] , [col3], ...]
        
    startCol:   int
                index of staring column \n
                if 0, the next colume of the last column of the excel file
        
    startRow:   int
                index of staring row \n
                if 0, the next row of the last column of the excel file \n
                1, the first row
    """
    
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name(sheetname)
    
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook[sheetname]
    
    if startCol == 0:
        indCol = ws.ncols+1
    else:
        indCol = startCol
    
    #column
    for lst in insert:
        #cell
        if startRow == 0:
            indRow = ws.nrows+1
        else:
            indRow = startRow
        for attr in lst:
            try:
                worksheet.cell(row= indRow, column = indCol).value = attr
            except(TypeError):
                print ('Type Error - '+str(indCol))
    
            indRow+=1
        indCol+=1
    workbook.save(filepath)
    print('saved successfully in existing file!') 

def write_row(filepath, sheetname, insert, startCol=1, startRow=0):
    """
    Method for adding a set of rows in an existing excel file
    
    Parameters
    ----------
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
        
    insert:     nested list
                [[row1], [row2], [row3], ...]
        
    startCol:   int
                index of staring column \n
                if 0, the next colume of the last column of the excel file \n
                1, A column of excel file
        
    startRow:   int
                index of staring row \n
                if 0, the next row of the last column of the excel file \n
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name(sheetname)
    
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook[sheetname]
#    worksheet = workbook.get_sheet_by_name(sheetname)
    
    if startRow == 0:
        indRow = ws.nrows+1
    else:
        indRow = startRow
    #row
    for lst in insert:
        #cell
        if startCol == 0:
            indCol = ws.ncols+1
        else:
            indCol = startCol
        for attr in lst:
            try:
                worksheet.cell(row= indRow, column = indCol).value = attr
            except(TypeError):
                print ('Type Error - '+str(indCol))
            
            indCol+=1
        indRow+=1
    workbook.save(filepath)
    print('saved successfully in existing file!')
    
def pointGraph_to_xlsx(pointGraph, filepath, sheetname):
    """
    Method for converting a point graph to an excel file
    
    Parameters
    ----------
    pointGraph:     networkx graph  
                    point type networkx graph to convert
                    
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
    """
    colNames = 'nodeX nodeY'.split(sep=' ')
    for col in pointGraph.nodes(data=True)[0][1].keys():
        colNames.append(col)
        
    dataList = []
    for point in pointGraph.nodes(data=True):
        rowList = [point[0][0], point[0][1]]
        for val in point[1].values():
            rowList.append(val)
        dataList.append(rowList)
    
    dataList.insert(0, colNames)
    
    try:
        clear_sheet(filepath, sheetname)
    except FileNotFoundError:
        create_xlsx(filepath, sheetname)
    finally:
        write_row(filepath, sheetname, dataList)
        
def nestedDict_to_xlsx(nestedDict, filepath, sheetname, colNames=[]):
    """
    Method for converting nested dictionary to excel file
    
    Parameters
    ----------
    nestedDict:     nested dictionary
                    dicationry to convert to an excel file
                    
    filepath:   string
                path of the excel file to handle            
    
    sheetname:  string
                name of sheet to access
                
    colNames:   list of string
                a list of column names. The number of column names should be same to that of data variables
    """
    survivedType = ["<class 'str'>", "<class 'float'>" , "<class 'int'>", "<class 'bool'>"]
    if colNames == []:
        columns = [str(key) for key in nestedDict.values()[0].keys()]
        columns.insert(0, 'key')
    else:
        assert(len(colNames) == len(list(nestedDict.values())[0].keys()) + 1 )
        
    dataList = [colNames]
    for key1, val1 in nestedDict.items():
        rowList = [key1]
        for key2, val2 in val1.items():
            if str(type(val2)) in survivedType: # only acceptable type takes its own type
                rowList.append(val2)
            else: # otherwise, it takes str type
                rowList.append(str(val2))
        dataList.append(rowList)
        
    try:
        clear_sheet(filepath, sheetname)
    except FileNotFoundError:
        create_xlsx(filepath, sheetname)
    finally:
        write_row(filepath, sheetname, dataList)
    
        
    
        